"""Core sync logic — single merged table (articles + GSC metrics)."""

import requests
from datetime import datetime, timedelta, UTC
from urllib.parse import urlparse

from src.config import (
    WP_API_BASE, GSC_CREDENTIALS_FILE, GSC_CREDENTIALS_JSON_B64,
    GSC_SITE_URL, HTTPS_PROXY, FEISHU_TABLE_ARTICLES, FEISHU_TABLE_HISTORY,
    FEISHU_BITABLE_APP_TOKEN,
)
from src.feishu_client import FeishuClient

BASE_FS = "https://open.feishu.cn/open-apis"


# ---- WordPress ----

def _fetch_all_wp_posts():
    posts = []
    page = 1
    while True:
        r = requests.get(f"{WP_API_BASE}/posts", params={
            "per_page": 100, "page": page,
            "_fields": "id,slug,title,date,link,categories,tags",
        })
        if r.status_code == 400:
            break
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        posts.extend(batch)
        page += 1
    return posts


def _fetch_recent_wp_posts(days=3):
    after = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00")
    posts = []
    page = 1
    while True:
        r = requests.get(f"{WP_API_BASE}/posts", params={
            "per_page": 100, "page": page, "after": after,
            "_fields": "id,slug,title,date,link,categories,tags",
        })
        if r.status_code == 400:
            break
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        posts.extend(batch)
        page += 1
    return posts


# ---- GSC ----

def _get_gsc_service():
    import httplib2
    import socks
    import json
    import base64
    from google.oauth2 import service_account
    from google_auth_httplib2 import AuthorizedHttp
    from googleapiclient.discovery import build

    SCOPES = ["https://www.googleapis.com/auth/webmasters.readonly"]
    if GSC_CREDENTIALS_JSON_B64:
        info = json.loads(base64.b64decode(GSC_CREDENTIALS_JSON_B64))
        creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    else:
        creds = service_account.Credentials.from_service_account_file(GSC_CREDENTIALS_FILE, scopes=SCOPES)

    if HTTPS_PROXY:
        from urllib.parse import urlparse as _up
        p = _up(HTTPS_PROXY)
        httplib2.socks = socks
        pi = httplib2.ProxyInfo(socks.PROXY_TYPE_HTTP, p.hostname, p.port or 7890)
        h = httplib2.Http(proxy_info=pi)
        authed = AuthorizedHttp(creds, http=h)
        return build("searchconsole", "v1", http=authed)
    return build("searchconsole", "v1", credentials=creds)


def _fetch_gsc_data(end_date, days=90):
    """Fetch GSC per-page metrics aggregated over `days` days. Returns {slug: {...}}"""
    svc = _get_gsc_service()
    start_date = (datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=days)).strftime("%Y-%m-%d")
    print(f"[gsc] Date range: {start_date} → {end_date} ({days} days)")

    resp = svc.searchanalytics().query(siteUrl=GSC_SITE_URL, body={
        "startDate": start_date, "endDate": end_date,
        "dimensions": ["page"], "rowLimit": 5000,
    }).execute()
    page_rows = {r["keys"][0]: r for r in resp.get("rows", [])}

    # Aggregate all language versions into one slug
    slug_agg = {}
    for url, row in page_rows.items():
        path = urlparse(url).path.strip("/")
        slug = path.split("/")[-1] if path else ""
        if not slug:
            continue
        impr = row.get("impressions", 0)
        clicks = row.get("clicks", 0)
        pos = row.get("position", 0)

        if slug not in slug_agg:
            slug_agg[slug] = {"impressions": 0, "clicks": 0, "pos_sum": 0}
        slug_agg[slug]["impressions"] += impr
        slug_agg[slug]["clicks"] += clicks
        slug_agg[slug]["pos_sum"] += pos * impr

    result = {}
    for slug, agg in slug_agg.items():
        impr = agg["impressions"]
        clicks = agg["clicks"]
        ctr = (clicks / impr * 100) if impr > 0 else 0
        avg_pos = (agg["pos_sum"] / impr) if impr > 0 else 0
        result[slug] = {
            "impressions": impr,
            "clicks": clicks,
            "ctr": round(ctr, 2),
            "avg_position": round(avg_pos, 1),
        }
    return result


# ---- Feishu field management ----

def _ensure_fixed_gsc_fields(fs):
    """Ensure fixed GSC columns exist (overwritten each week)."""
    headers = {"Authorization": f"Bearer {fs.token}"}
    r = requests.get(
        f"{BASE_FS}/bitable/v1/apps/{FEISHU_BITABLE_APP_TOKEN}/tables/{FEISHU_TABLE_ARTICLES}/fields",
        headers=headers,
    )
    existing = {f["field_name"] for f in r.json()["data"]["items"]}

    needed = [
        ("impressions_3m", 2),
        ("clicks_3m", 2),
        ("ctr_3m", 2),
        ("avg_position_3m", 2),
        ("gsc_updated", 1),  # text: date of last GSC update
    ]
    for name, ftype in needed:
        if name not in existing:
            requests.post(
                f"{BASE_FS}/bitable/v1/apps/{FEISHU_BITABLE_APP_TOKEN}/tables/{FEISHU_TABLE_ARTICLES}/fields",
                headers=headers,
                json={"field_name": name, "type": ftype},
            )
            print(f"  Added field: {name}")


# ---- Public API ----

def sync_new_wp_articles(days=3):
    """Add new WordPress articles to Feishu (incremental)."""
    posts = _fetch_recent_wp_posts(days)
    if not posts:
        print("[wp] No new posts")
        return 0

    fs = FeishuClient()
    existing = fs.list_records(FEISHU_TABLE_ARTICLES)
    existing_slugs = {r["fields"].get("slug") for r in existing if r["fields"].get("slug")}

    new_records = []
    for p in posts:
        slug = p["slug"]
        if slug in existing_slugs:
            continue
        new_records.append({
            "slug": slug,
            "url": p["link"],
            "title": p["title"]["rendered"],
            "publish_time": p["date"][:10],
            "status": "新文章",
            "last_updated": datetime.now(UTC).strftime("%Y-%m-%d"),
        })

    if new_records:
        fs.batch_create(FEISHU_TABLE_ARTICLES, new_records)
    print(f"[wp] {len(new_records)} new, {len(posts) - len(new_records)} skipped")
    return len(new_records)


def update_gsc_data(date_str=None):
    """Fetch GSC 3-month data and overwrite fixed columns in article table."""
    if not date_str:
        date_str = (datetime.now(UTC) - timedelta(days=2)).strftime("%Y-%m-%d")

    fs = FeishuClient()
    _ensure_fixed_gsc_fields(fs)

    print(f"[gsc] Fetching 3-month data ending {date_str}...")
    gsc = _fetch_gsc_data(date_str, days=90)
    print(f"[gsc] {len(gsc)} pages with data")

    records = fs.list_records(FEISHU_TABLE_ARTICLES)
    slug_to_rid = {r["fields"].get("slug"): r["record_id"] for r in records if r["fields"].get("slug")}

    updates = []
    for slug, data in gsc.items():
        rid = slug_to_rid.get(slug)
        if not rid:
            continue
        fields = {
            "impressions_3m": data["impressions"],
            "clicks_3m": data["clicks"],
            "ctr_3m": data["ctr"],
            "avg_position_3m": data["avg_position"],
            "gsc_updated": date_str,
        }
        updates.append((rid, fields))

    if updates:
        fs.batch_update(FEISHU_TABLE_ARTICLES, updates)
    print(f"[gsc] Updated {len(updates)} articles (3-month data as of {date_str})")
    return len(updates)


def save_weekly_snapshot(date_str=None):
    """Save 7-day GSC data to history table for trend tracking."""
    if not date_str:
        date_str = (datetime.now(UTC) - timedelta(days=2)).strftime("%Y-%m-%d")

    print(f"[snapshot] Fetching 7-day data ending {date_str}...")
    gsc_7d = _fetch_gsc_data(date_str, days=7)
    print(f"[snapshot] {len(gsc_7d)} pages with data")

    # Get article info from main table (only articles published in last year)
    one_year_ago = (datetime.now(UTC) - timedelta(days=365)).strftime("%Y-%m-%d")
    fs = FeishuClient()
    records = fs.list_records(FEISHU_TABLE_ARTICLES)
    slug_info = {}
    for rec in records:
        f = rec["fields"]
        slug = f.get("slug", "")
        pub = f.get("publish_time", "")
        if slug and pub and pub >= one_year_ago:
            slug_info[slug] = {
                "title": f.get("title", ""),
                "author": f.get("author", ""),
                "category": f.get("category", ""),
            }

    # Build snapshot records (all articles with GSC data, published in last year)
    snapshot = []
    for slug, data in gsc_7d.items():
        info = slug_info.get(slug, {})
        snapshot.append({
            "week_date": date_str,
            "slug": slug,
            "title": info.get("title", slug),
            "author": info.get("author", ""),
            "category": info.get("category", ""),
            "clicks_7d": data["clicks"],
            "impressions_7d": data["impressions"],
            "ctr_7d": data["ctr"],
            "avg_position_7d": data["avg_position"],
        })

    if snapshot:
        fs.batch_create(FEISHU_TABLE_HISTORY, snapshot)
    print(f"[snapshot] Saved {len(snapshot)} articles to history for {date_str}")
    return len(snapshot)


def sync_author():
    """Sync author=阿水 from Feishu tracking table + Excel."""
    import re
    fs = FeishuClient()
    records = fs.list_records(FEISHU_TABLE_ARTICLES)
    slug_to_rid = {}
    already_marked = set()
    for r in records:
        f = r["fields"]
        slug = f.get("slug", "")
        if slug:
            slug_to_rid[slug] = r["record_id"]
            if f.get("author") == "阿水":
                already_marked.add(slug)

    ashuia_slugs = set()

    # Source 1: Feishu tracking table (WP post ID → slug)
    AUTH_APP_TOKEN = "C1wLbnYWIaAtzksJU8xcUKMyn8g"
    AUTH_TABLE_ID = "tblGffwKgoXem01b"
    auth_records = []
    page_token = None
    while True:
        params = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        r = requests.get(
            f"{BASE_FS}/bitable/v1/apps/{AUTH_APP_TOKEN}/tables/{AUTH_TABLE_ID}/records",
            headers={"Authorization": f"Bearer {fs.token}"}, params=params,
        )
        data = r.json()["data"]
        auth_records.extend(data.get("items") or [])
        if not data.get("has_more"):
            break
        page_token = data["page_token"]

    # Extract WP post IDs → slugs
    post_ids = set()
    for rec in auth_records:
        f = rec["fields"]
        link = f.get("文章wordpress链接", "") or ""
        m = re.search(r'[?&](?:post|p)=(\d+)', link)
        if m:
            post_ids.add(m.group(1))

    if post_ids:
        pid_list = list(post_ids)
        for i in range(0, len(pid_list), 100):
            batch = pid_list[i:i+100]
            r = requests.get(f"{WP_API_BASE}/posts", params={
                "include": ",".join(batch), "per_page": 100, "_fields": "id,slug",
            })
            if r.status_code == 200:
                for p in r.json():
                    ashuia_slugs.add(p["slug"])

    # Update missing
    updates = []
    for slug in ashuia_slugs:
        if slug not in already_marked and slug in slug_to_rid:
            updates.append((slug_to_rid[slug], {"author": "阿水"}))

    if updates:
        fs.batch_update(FEISHU_TABLE_ARTICLES, updates)
    print(f"[author] {len(already_marked)} existing + {len(updates)} new = {len(already_marked) + len(updates)} total 阿水 articles")


def update_status():
    """Auto-tag articles based on latest GSC data."""
    fs = FeishuClient()
    records = fs.list_records(FEISHU_TABLE_ARTICLES)

    updates = []
    for rec in records:
        f = rec["fields"]
        rid = rec["record_id"]
        impr = _num(f.get("impressions_3m", 0))
        clicks = _num(f.get("clicks_3m", 0))
        ctr = _num(f.get("ctr_3m", 0))
        pos = _num(f.get("avg_position_3m", 0))

        if impr == 0 and clicks == 0:
            status = "无数据"
        elif impr < 10:
            status = "收录期"
        elif impr > 500 and ctr < 1.0:
            status = "CTR异常"
        elif pos > 0 and pos <= 10:
            status = "首页"
        else:
            status = "正常"

        if status != f.get("status", ""):
            updates.append((rid, {"status": status}))

    if updates:
        fs.batch_update(FEISHU_TABLE_ARTICLES, updates)
    print(f"[status] Updated {len(updates)} articles")


def full_init():
    """First-time full import: all WP posts + categories + GSC data."""
    print("=== Full init ===")
    posts = _fetch_all_wp_posts()
    print(f"[init] WP: {len(posts)} posts")

    slug_cat = _load_categories()
    print(f"[init] Categories: {len(slug_cat)} mappings")

    date_str = (datetime.now(UTC) - timedelta(days=2)).strftime("%Y-%m-%d")
    print(f"[init] Fetching GSC for {date_str}...")
    gsc = _fetch_gsc_data(date_str)
    print(f"[init] GSC: {len(gsc)} pages")

    fs = FeishuClient()
    _ensure_fixed_gsc_fields(fs)

    records = []
    for p in posts:
        slug = p["slug"]
        rec = {
            "slug": slug,
            "url": p["link"],
            "title": p["title"]["rendered"],
            "publish_time": p["date"][:10],
            "status": "正常",
            "last_updated": date_str,
            "category": slug_cat.get(slug, ""),
        }
        g = gsc.get(slug, {})
        if g:
            rec["impressions_3m"] = g["impressions"]
            rec["clicks_3m"] = g["clicks"]
            rec["ctr_3m"] = g["ctr"]
            rec["avg_position_3m"] = g["avg_position"]
            rec["gsc_updated"] = date_str

        impr = g.get("impressions", 0)
        ctr_val = g.get("ctr", 0)
        if impr == 0:
            rec["status"] = "无数据"
        elif impr < 10:
            rec["status"] = "收录期"
        elif impr > 500 and ctr_val < 1.0:
            rec["status"] = "CTR异常"
        else:
            rec["status"] = "正常"

        records.append(rec)

    fs.batch_create(FEISHU_TABLE_ARTICLES, records)
    with_gsc = sum(1 for r in records if r.get("impressions_3m", 0) > 0)
    print(f"[init] Inserted {len(records)} articles ({with_gsc} with GSC data)")


def _load_categories():
    """Load slug→category from Excel file."""
    from src.config import CATEGORY_EXCEL
    if not CATEGORY_EXCEL:
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(CATEGORY_EXCEL, read_only=True)
        ws = wb[wb.sheetnames[0]]
        result = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            slug = str(row[5]).strip() if row[5] else ""
            cat = str(row[2]).strip() if row[2] else ""
            if slug and cat:
                result[slug] = cat
        return result
    except Exception as e:
        print(f"[categories] Warning: {e}")
        return {}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0
